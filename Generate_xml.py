import datetime
import openpyxl
import xml.dom.minidom as md
from xml.etree.ElementTree import Element, SubElement, tostring

def add_attribute(root):
    root.set('xmlns:gml','http://www.opengis.net/gml')
    root.set('xmlns:xlink','http://www.w3.org/1999/xlink')
    root.set('xmlns:xsi','http://www.w3.org/2001/XMLSchema-instance')
    root.set('xmlns:gco','http://www.isotc211.org/2005/gco')
    root.set('xmlns:gmd','http://www.isotc211.org/2005/gmd')
    root.set('xsi:schemaLocation','https://standards.moi.gov.tw/schema/utilityex utilityex.xsd')
    root.set('xmlns','https://standards.moi.gov.tw/schema/utilityex')
    return

def add_TimeInstant(sheet,col_index,element,value):
    if(sheet.cell(row = 2, column = col_index).value == '設置日期'):
        Time_element = SubElement(element,'gml:TimeInstant')
        Time_pos = SubElement(Time_element,'gml:timePosition')
        date_value = datetime.datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        formatted_date = date_value.strftime('%Y-%m-%d')
        Time_pos.text = formatted_date
        return
    else:
        element.text = value
        return

def xlsx_to_xml(input_file):
    wb = openpyxl.load_workbook(input_file)
    
    
    for index in range(len(wb.sheetnames)):
        
        root = Element('UTL')
        add_attribute(root)
        
        sheet_name = wb.sheetnames[index] #取得表名
        
        
        
        
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        for row_index in range(2, max_row ):
            gml_element = SubElement(root,"gml:featureMember") #root子元素 gml:featureMember
            safe_sheet_name = sheet_name.replace("+", "_") #替換非法字元
            sheet_element = SubElement(gml_element,safe_sheet_name) #gml:featureMember子元素 表名
            geometry_title = SubElement(sheet_element,'geometry')
            geometry_element = SubElement(geometry_title,'gml:Point', srsName="EPSG:3826",srsDimension="3")
            geometry_coord = SubElement(geometry_element,'gml:coordinates')
            for column_index in range(1, max_column):
                cell = sheet.cell(row = row_index+1, column = column_index)
                cell_value = str(cell.value) if cell.value is not None else ""
                geometry_coord.text = str(sheet.cell(row = row_index+1, column = max_column).value)  
                cell_element = SubElement(sheet_element, str(sheet.cell(row = 2, column = column_index).value))
                add_TimeInstant(sheet, column_index, cell_element, cell_value)
               # cell_element.text = cell_value
    
        xml_str = tostring(root, encoding='utf-8')
        xml_dom = md.parseString(xml_str)
        #formatted_xml = xml_dom.toprettyxml(indent='\t')
        #print(formatted_xml)
    
        output_file = f'{sheet_name}.gml'
    
        with open(output_file, 'w', encoding = 'utf-8') as file:
            xml_dom.writexml(file, indent = '\t', addindent = '\t', newl = '\n', encoding = 'utf-8')


input_file = "pos1.xlsx"
input_file_2 = "pos2.xlsx"
xlsx_to_xml(input_file)
xlsx_to_xml(input_file_2)
